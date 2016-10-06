import re
import openpyxl as xl

import bookcase_exceptions as exc
import bookcase_lib as lib
from bookcase_translations import Translations


class Excel(object):
    """
    Class for input/output to Excel worksheet
    """
    def __init__(self, path=lib.FileManager().path):
        self.path = path

    @staticmethod
    def validate_filename(filename):
        """
        Make sure that filename is consistent with the format of Bookcase manager
        generated Excel files
        :param filename:
        :return: name of the database exported
        """
        exp = re.compile(r'(?<=bookcase_).*(?=\.xlsx)')
        name = re.search(exp, filename)
        if not name:
            raise exc.InvalidInputException(Translations().invalid_xl_msg)
        return name

    def read_excel(self, filename):
        """
        Reads the table from the specified Excel file
        :param filename:
        :return: tuple of tuples representing the table
        """
        self.validate_filename(filename)
        workbook = xl.load_workbook(self.path + filename)
        result = []
        for row in workbook.active.iter_rows():
            result.append(tuple(cell.value for cell in row))
        return tuple(result)

    def write_excel(self, name, table):
        """
        Writes a table to an excel file
        :param name: the name of the exported DB
        :param table: 2-D array containing DB table dump
        :return: absolute path to file + filename
        """
        workbook = xl.Workbook()
        worksheet = workbook.active
        for row_num, row in enumerate(table):
            for col_num, value in enumerate(row):
                worksheet.cell(row=row_num + 1, column=col_num + 1, value=value)
        file = self.path + "bookcase_{name}.xlsx".format(name=name)
        workbook.save(file)
        return file
